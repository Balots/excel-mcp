import logging
import os
from pathlib import Path
from typing import Any
import requests

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .exceptions import WorkbookError

logger = logging.getLogger(__name__)

# 静态资源服务器配置
FILE_SERVER_URL = "http://localhost:3001"
UPLOAD_ENDPOINT = f"{FILE_SERVER_URL}/upload"
FILES_LIST_ENDPOINT = f"{FILE_SERVER_URL}/files/list"
FILE_ACCESS_BASE_URL = f"{FILE_SERVER_URL}/files/"

def upload_file_to_server(filepath: str) -> dict[str, Any]:
    """Upload a file to the static file server
    
    Args:
        filepath: Path to the file to upload
        
    Returns:
        Dictionary with upload result information including file URL
    """
    try:
        if not os.path.exists(filepath):
            raise WorkbookError(f"File not found: {filepath}")
            
        filename = os.path.basename(filepath)
        
        with open(filepath, 'rb') as file:
            files = {'file': (filename, file)}
            response = requests.post(UPLOAD_ENDPOINT, files=files)
            
        if response.status_code != 200:
            raise WorkbookError(f"Failed to upload file: {response.text}")
            
        # 构建文件访问URL
        file_url = f"{FILE_ACCESS_BASE_URL}{filename}"
        
        return {
            "message": f"File uploaded successfully",
            "file_url": file_url,
            "filename": filename
        }
    except Exception as e:
        logger.error(f"Failed to upload file: {e}")
        raise WorkbookError(f"Failed to upload file: {e!s}")

def download_file_from_url(url: str, save_path: str) -> str:
    """Download a file from URL and save it to the specified path
    
    Args:
        url: URL of the file to download
        save_path: Path where to save the downloaded file
        
    Returns:
        Path to the downloaded file
    """
    try:
        response = requests.get(url, stream=True)
        response.raise_for_status()
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        
        with open(save_path, 'wb') as file:
            for chunk in response.iter_content(chunk_size=8192):
                file.write(chunk)
                
        return save_path
    except Exception as e:
        logger.error(f"Failed to download file: {e}")
        raise WorkbookError(f"Failed to download file: {e!s}")

def create_workbook(filepath: str, sheet_name: str = "Sheet1", upload: bool = False) -> dict[str, Any]:
    """Create a new Excel workbook with optional custom sheet name and upload to server"""
    try:
        wb = Workbook()
        # Rename default sheet
        if "Sheet" in wb.sheetnames:
            sheet = wb["Sheet"]
            sheet.title = sheet_name
        else:
            wb.create_sheet(sheet_name)

        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        
        result = {
            "message": f"Created workbook: {filepath}",
            "active_sheet": sheet_name,
            "workbook": wb
        }
        
        # 如果需要上传文件
        if upload:
            upload_result = upload_file_to_server(filepath)
            result["file_url"] = upload_result["file_url"]
            result["message"] = f"Created and uploaded workbook: {filepath}. URL: {upload_result['file_url']}"
            
        return result
    except Exception as e:
        logger.error(f"Failed to create workbook: {e}")
        raise WorkbookError(f"Failed to create workbook: {e!s}")

def get_or_create_workbook(filepath: str) -> Workbook:
    """Get existing workbook or create new one if it doesn't exist"""
    try:
        return load_workbook(filepath)
    except FileNotFoundError:
        return create_workbook(filepath)["workbook"]

def create_sheet(filepath: str, sheet_name: str) -> dict:
    """Create a new worksheet in the workbook if it doesn't exist."""
    try:
        wb = load_workbook(filepath)

        # Check if sheet already exists
        if sheet_name in wb.sheetnames:
            raise WorkbookError(f"Sheet {sheet_name} already exists")

        # Create new sheet
        wb.create_sheet(sheet_name)
        wb.save(filepath)
        wb.close()
        return {"message": f"Sheet {sheet_name} created successfully"}
    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to create sheet: {e}")
        raise WorkbookError(str(e))

def get_workbook_info(filepath: str, include_ranges: bool = False) -> dict[str, Any]:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        path = Path(filepath)
        if not path.exists():
            raise WorkbookError(f"File not found: {filepath}")
            
        wb = load_workbook(filepath, read_only=True)
        
        info = {
            "filename": path.name,
            "sheets": wb.sheetnames,
            "size": path.stat().st_size,
            "modified": path.stat().st_mtime
        }
        
        if include_ranges:
            # Add used ranges for each sheet
            ranges = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row > 0 and ws.max_column > 0:
                    ranges[sheet_name] = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            info["used_ranges"] = ranges
            
        wb.close()
        return info
        
    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to get workbook info: {e}")
        raise WorkbookError(str(e))
